from google.cloud import storage
from jinja2 import Environment, FileSystemLoader
from io import BytesIO

import json
import os
import openpyxl
import excelProcessing as ep

'''
TODO MAIN:
- make a log file
- look at where to error handle

TODO NOW:
- save the templates to the website directory (this is done in the render functions but move it out to make for easier testing)
'''


class ScenarioTool:
    class MissingProvenanceError(Exception):
        def __init__(self, message, project):
            message  = f"Provenance is missing. Please upload a provenance file to {project}/standards/provenance.json"
            self.message = message
            super().__init__(self.message)
    
    class MissingContinuousDataError(Exception):
        def __init__(self, message, project, personae, sheet):
            message = f"Continuous data is missing. Please upload a continuous data file to {project}/continuous_data/{personae}-{sheet}.csv"
            self.message = message
            super().__init__(self.message)
            
    class MissingStandardError(Exception):
        def __init__(self, message, project):
            message = f"Standard is missing. Please upload a standard file to {project}/standards/"
            self.message = message
            super().__init__(self.message)
            
    class MissingPersonaeError(Exception):
        def __init__(self, message, project):
            message = f"Personae is missing. Please upload a personae file to {project}/personae/"
            self.message = message
            super().__init__(self.message)
            
    class CompileError(Exception):
        def __init__(self, message, project):
            message = f"Compile error. {project} check the logs for more information"
            self.message = message
            super().__init__(self.message)

    def __init__(self, project_name, gcp_site):
        self.project_name = project_name.lower()
        self.gcp_project = gcp_site
        self.validated = False
        self.standard_dir = f'{gcp_site}/{project_name}/standards'
        self.personae_dir = f'{gcp_site}/{project_name}/personae'
        self.continuous_data_dir = f'{gcp_site}/{project_name}/continuous_data'
        self.storage_client = storage.Client()
        self.provenance = []
        
        # initiate a template environment
        this_file_dir = os.path.dirname(os.path.abspath(__file__))
        parent_dir = this_file_dir[:this_file_dir.rfind('/')]
        tempalte_path = os.path.join(parent_dir, 'templates')
        self.template_env = Environment(loader=FileSystemLoader(tempalte_path))
        
        self.bucket = self.storage_client.bucket(self.gcp_project)
        
    def copy_project(self, project_name:str, new_project_name: str) -> None:
        blobs = self.bucket.list_blobs(prefix=project_name)
        for blob in blobs:
            # Construct the destination object name
            destination_blob_name = blob.name.replace(project_name, new_project_name, 1)

            # Copy the blob to the new location
            self.bucket.copy_blob(blob, self.bucket, destination_blob_name)
        self.project_name = project_name
        
        return
    
    def list_projects(self) -> list:
        blobs = self.bucket.list_blobs(prefix=self.gcp_project)
        folders = set()
        for blob in blobs:
            folder = blob.name.split('/')[1]      
            folders.add(folder)
        
        return list(folders)
    
    def set_project(self, new_project_name: str) -> None:
        self.project_name = new_project_name
        self.standard_dir = f'{self.gcp_project}/{new_project_name}/standards'
        self.personae_dir = f'{self.gcp_project}/{new_project_name}/personae'
        self.continuous_data_dir = f'{self.gcp_project}/{new_project_name}/continuous_data'
        
        return
    
    def del_project(self) -> None:
        # Delete all blobs in the project folder
        blobs = self.bucket.list_blobs(prefix=f'{self.gcp_project}/{self.project_name}')
        for blob in blobs:
            blob.delete()

        # Delete the project folder
        self.bucket.delete_blob(f'{self.gcp_project}/{self.project_name}')
        print(f"Project {self.project_name} deleted")

        # Reset project attributes
        self.project_name = None
        self.standard_dir = None
        self.personae_dir = None
        self.continuous_data_dir = None
        self.validated = False
        self.provenance = []

        return
        
    def upload_provenance(self, file_path: str) -> None:
        '''
        A function to upload a provenance file to the personae directory
        Re-create all path lists
        Re-validate personae
        
        :param file_path: The path to the provenance file to upload
        :return: None
        '''
        # get path list
        provenance_paths = ep.get_standard_paths(file_path, [], False)
        provenance_paths_json_data = json.dumps(provenance_paths)
        self.provenance = provenance_paths
        
        # upload json string to provenance.json file
        blob = self.bucket.blob(f'{self.standard_dir}/provenance.json')
        blob.upload_from_string(provenance_paths_json_data)
        print(f"File {file_path} uploaded to {self.standard_dir}")
        
        # make path list json
        standard_paths = ep.get_standard_paths(file_path, self.provenance, False)
        
        # save to the standarsds directory
        file_name = file_name.replace('.xlsx', '.json')
        standard_paths_json_data = json.dumps(standard_paths)
        blob = self.bucket.blob(f'{self.standard_dir}/{file_name}')
        blob.upload_from_string(standard_paths_json_data)
        
        # loop over personae and validate
        # Get the list of blobs in the personae directory
        blobs = self.bucket.list_blobs(prefix=self.personae_dir)

        # Loop over each blob in the personae directory
        for blob in blobs:
            
            # Extract the file name from the blob
            file_name = blob.name.split('/')[-1]
            false_paths = self.validate_personae(file_name)
            
            # save to a json file
            false_paths_json_data = json.dumps(false_paths)
            blob = self.bucket.blob(f'{self.personae_dir}/{file_name.replace(".xlsx", ".falsepaths.json")}')
            blob.upload_from_string(false_paths_json_data)

        return
    
    def upload_continuous_data(self, file_path: str) -> None:
        '''
        A function to upload a continuous data file to the personae directory
        
        :param file_path: The path to the continuous data file to upload
        :return: None
        '''
        file_name = os.path.basename(file_path)
        blob = self.bucket.blob(f'{self.continuous_data_dir}/{file_name}')
        blob.upload_from_filename(file_path)
        print(f"File {file_path} uploaded to {self.continuous_data_dir}")
        
        return
    
    def upload_patient(self, file_path: str) -> None:
        '''
        A function to upload a patient storylog to the personae directory
        Re-validate personae
        
        :param file_path: The path to the patient storylog to upload
        :return: None
        '''
        # upload file
        file_name = os.path.basename(file_path)
        blob = self.bucket.blob(f'{self.personae_dir}/{file_name}')
        blob.upload_from_filename(file_path)
        print(f"File {file_path} uploaded to {self.personae_dir}")
        
        # loop over personae and validate
        # Get the list of blobs in the personae directory
        blobs = self.bucket.list_blobs(prefix=self.personae_dir)

        # Loop over each blob in the personae directory
        for blob in blobs:
            
            # Extract the file name from the blob
            file_name = blob.name.split('/')[-1]
            false_paths, storylog = self.validate_personae(file_name)
            
            # save to a json file
            false_paths_json_data = json.dumps(false_paths)
            blob = self.bucket.blob(f'{self.personae_dir}/{file_name.replace(".xlsx", ".falsepaths.json")}')
            blob.upload_from_string(false_paths_json_data)
            
            # save storylog
            storylog_json_data = json.dumps(storylog)
            blob = self.bucket.blob(f'{self.personae_dir}/{file_name.replace(".xlsx", ".storylog.json")}')
            blob.upload_from_string(storylog_json_data)
            
        print("Personae validated, storylogs created")
        
        return

    def upload_standard(self, file_path: str) -> None:
        '''
        A function to upload a standard to the standards directory
        Creates path list for standard
        Re-validate personae
        
        :param file_path: The path to the standard xlsx file to upload
        :return: None
        '''
        if self.provenance == []:
            raise self.MissingProvenanceError(self.provenance, self.project_name)
        
        # upload file
        file_name = os.path.basename(file_path)
        blob = self.bucket.blob(f'{self.standard_dir}/{file_name}')
        blob.upload_from_filename(file_path)
        
        # make path list json
        standard_paths = ep.get_standard_paths(file_path, self.provenance, False)
        
        # save to the standarsds directory
        file_name = file_name.replace('.xlsx', '.json')
        standard_paths_json_data = json.dumps(standard_paths)
        blob = self.bucket.blob(f'{self.standard_dir}/{file_name}')
        blob.upload_from_string(standard_paths_json_data)
        
        # loop over personae and validate
        # Get the list of blobs in the personae directory
        blobs = self.bucket.list_blobs(prefix=self.personae_dir)

        # Loop over each blob in the personae directory
        for blob in blobs:
            
            # Extract the file name from the blob
            file_name = blob.name.split('/')[-1]
            false_paths, storylog = self.validate_personae(file_name)
            
            # save to a json file
            false_paths_json_data = json.dumps(false_paths)
            blob = self.bucket.blob(f'{self.personae_dir}/{file_name.replace(".xlsx", ".falsepaths.json")}')
            blob.upload_from_string(false_paths_json_data)
            
            # save storylog
            storylog_json_data = json.dumps(storylog)
            blob = self.bucket.blob(f'{self.personae_dir}/{file_name.replace(".xlsx", ".storylog.json")}')
            blob.upload_from_string(storylog_json_data)
            
        print(f"File {file_path} uploaded to {self.standard_dir}")
        print("Personae validated")
        
        return
    
    def del_standard(self, standard_file_name: str) -> None:
        '''
        A function to delete a standard from the standards directory
        Removes path list for standard
        
        :param standard_file_name: The name of the standard file to delete
        :return: None
        '''
        blob = self.bucket.blob(f'{self.standard_dir}/{standard_file_name}')
        blob.delete()
        print(f"File {standard_file_name} deleted from {self.standard_dir}")
        
        # loop over personae and validate
        # Get the list of blobs in the personae directory
        blobs = self.bucket.list_blobs(prefix=self.personae_dir)

        # Loop over each blob in the personae directory
        for blob in blobs:
            
            # Extract the file name from the blob
            file_name = blob.name.split('/')[-1]
            false_paths, storylog = self.validate_personae(file_name)
            
            # save to a json file
            false_paths_json_data = json.dumps(false_paths)
            blob = self.bucket.blob(f'{self.personae_dir}/{file_name.replace(".xlsx", ".falsepaths.json")}')
            blob.upload_from_string(false_paths_json_data)

            # save storylog
            storylog_json_data = json.dumps(storylog)
            blob = self.bucket.blob(f'{self.personae_dir}/{file_name.replace(".xlsx", ".storylog.json")}')
            blob.upload_from_string(storylog_json_data)

        print("Personae validated")
        
        return
    
    def standard_path_list(self, standard_name: str) -> list:
        '''
        A function to return the path list for a given standard
        
        :param standard_name: The name of the standard to return the path list for without extention
        :return: The path list for the given standard
        '''
        # get path list
        blob = self.bucket.blob(f'{self.standard_dir}/{standard_name}.json')
        standard_data_str = blob.download_as_string()
        standard_list = json.loads(standard_data_str)
        
        return standard_list
        
        
    def standards_list(self) -> list:
        '''
        A function to return a list of standards in the standards directory
        
        :return: A list of standards in the standards directory
        '''
        standard_fles = [blob.name for blob in self.bucket.list_blobs(prefix=f'{self.project_name}/standards/') if blob.name.endswith('.xlsx')]
        
        return standard_fles
    
    def patient_list(self) -> list:
        '''
        A function to return a list of patients in the personae directory
        
        :return: A list of patients in the personae directory
        '''
        standard_fles = [blob.name for blob in self.bucket.list_blobs(prefix=f'{self.project_name}/personae/')]
        
        return standard_fles
    
    def get_patient(self, patient_name: str) -> dict:
        '''
        A function to return a patient storylog from the personae directory
        
        :param patient_name: The name of the patient to return the storylog for
        :return: The storylog for the given patient
        '''
        # get patient
        try:
            blob = self.bucket.blob(f'{self.personae_dir}/{patient_name}.json')
        except:
            blob = self.bucket.blob(f'{self.personae_dir}/{patient_name}.xlsx')
            
        patient_str = blob.download_as_string()
        patient_json = json.loads(patient_str)

        return patient_json
    
    def save_patient(self, patient_file: dict, patient_name: str) -> None:
        '''
        A function to save a patient storylog to the personae directory
        Re-compiles website
        
        :param patient_file: The patient storylog to save
        :param patient_name: The name of the patient storylog to save
        :return: None
        '''
        # upload patient file
        patient_string = json.dumps(patient_file)
        blob = self.bucket.blob(f'{self.personae_dir}/{patient_name}.json')
        blob.upload_from_string(patient_string)
        print(f"File {patient_name}.json uploaded to {self.personae_dir}")
        
        # validate patient
        # loop over personae and validate
        # Get the list of blobs in the personae directory
        blobs = self.bucket.list_blobs(prefix=self.personae_dir)

        # Loop over each blob in the personae directory
        for blob in blobs:
            
            # Extract the file name from the blob
            file_name = blob.name.split('/')[-1]
            false_paths, storylog = self.validate_personae(file_name)
            
            # save to a json file
            false_paths_json_data = json.dumps(false_paths)
            blob = self.bucket.blob(f'{self.personae_dir}/{file_name.replace(".xlsx", ".falsepaths.json")}')
            blob.upload_from_string(false_paths_json_data)

            # save storylog
            storylog_json_data = json.dumps(storylog)
            blob = self.bucket.blob(f'{self.personae_dir}/{file_name.replace(".xlsx", ".storylog.json")}')
            blob.upload_from_string(storylog_json_data)

        print("Personae validated")
        
        # re-compile website
        self.compile_website()

        return
    
    def del_patient(self, patient_name: str) -> None:
        '''
        A function to delete a patient storylog from the personae directory
        Re-compiles website
        
        :param patient_name: The name of the patient storylog to delete
        :return: None
        '''
        # delete patient file
        blob = self.bucket.blob(f'{self.personae_dir}/{patient_name}.json')
        blob.delete()
        print(f"File {patient_name}.json deleted from {self.personae_dir}")
        
        # recompile website
        self.compile_website()
        
        return
    
    def validate_personae(self, personae_name: str) -> dict:
        '''
        A function to validate a personae against a given standard
        
        :param personae_name: The name of the personae to validate. Should include the extention
        :return: A dictionary of scenario sheets and their false paths
        '''
        false_paths = {}
        
        # open personae file
        blob = self.bucket.blob(f'{self.personae_dir}/{personae_name}')
        if personae_name.endswith('.json'):
            personae_data_str = blob.download_as_string()
            personae_data = json.loads(personae_data_str)
            
            # get the timeline items
            for event in personae_data['timeline']:
                event_standard = event['standard']
                event_linked_data = event['linked_data']
                false_paths['sheet'] = []
                
                # check if there is linked data
                if len(event_linked_data) != 0:
                    
                    # find the standard path list
                    try:
                        blob = self.bucket.blob(f'{self.standard_dir}/{event_standard}.json')
                    except:
                        raise self.MissingStandardError(event_standard, self.project_name)
                    
                    standard_data_str = blob.download_as_string()
                    standard_data = json.loads(standard_data_str)
                    
                    for item in event_linked_data:
                        if len(item['dataPath']) != 0:
                            if item['dataPath'] not in standard_data:
                                false_paths['sheet'].append(item['dataPath'])

        elif personae_name.endswith('.xlsx'):
            byte_stream = BytesIO()
            blob.download_to_file(byte_stream)
            byte_stream.seek(0)
            wb = openpyxl.load_workbook(filename=byte_stream, read_only=True)
            
            # initiate storylog
            storylog = {}
            
            # get the story sheet and work out what standard it is
            story_sheet = wb['Story']
            story_standard = str(story_sheet['E2'].value)
            
            # get timeline sheet
            timeline_sheet = wb['Time Line']
            
            storylog['summary'] = str(story_sheet['A2'].value)
            storylog['rationale'] = str(story_sheet['B2'].value)
            storylog['story'] = str(story_sheet['C2'].value)
            storylog['standard_url'] = str(story_sheet['D2'].value)
            storylog['standard_name'] = str(story_sheet['E2'].value)
            storylog['timeline'] = ep.get_timeline_storylog_format(timeline_sheet, story_standard)
            
            # get the standard path list
            try:
                blob = self.bucket.blob(f'{self.standard_dir}/{story_standard}.json')
            except:
                raise self.MissingStandardError(story_standard, self.project_name)
        
            standard_data_str = blob.download_as_string()
            standard_data = json.loads(standard_data_str)
            
            # get the scenario sheets
            scenario_sheets = [sheet for sheet in wb.sheetnames if sheet.lower() not in ['story', 'time line', 'timeline', 'group aliases', 'path aliases']]
            
            # get the group alias and path alias sheets
            group_alias_sheet = wb['Group Aliases']
            path_alias_sheet = wb['Path Aliases']
            
            # loop over the scenario sheets
            for sheet in scenario_sheets:
                false_paths[sheet] = []
                scenario_sheet = wb[sheet]
                
                path_location = ep.find_cell_location(scenario_sheet, 'Data Path')
                example_data_location = ep.find_cell_location(scenario_sheet, 'Example Data')
                
                path_list = []
                in_loop = False
                
                for row in scenario_sheet.iter_rows(min_row=path_location[0] + 1,
                                        max_row=scenario_sheet.max_row,
                                        min_col=path_location[1],
                                        max_col=example_data_location[1]):
                    
                    # loop over rows expanding syntax
                    if row[0].value:
                        path = row[0].value.lower().strip()
                        value = row[0].value
                        
                        # check if in a continuous data loop
                        if in_loop:
                            if path != '$loopend':
                                continue
                            elif path == '$loopend':
                                in_loop = False
                                continue
                            
                            # process group alias
                            elif '$$' in path:
                                group_alias_name = path.split('.')[-1].replace('$$', '').strip().lower()
                                without_alias = '.'.join(path.split('.')[:-1])
                                group_alias_paths = ep.expand_group_aliases(group_alias_sheet, group_alias_name)
                                
                                split_path = path.split('.')
                                for index, element in enumerate(split_path[:-1]):
                                    if element.startswith('$'):
                                        if '[' in element:
                                            element_no_index = element.split('[')[0]
                                            element_index = element.split('[')[1].replace(']', '')
                                            
                                            expanded_value = ep.expand_path_alias(element_no_index, path_alias_sheet)
                                            expanded_value = f'{expanded_value}[{element_index}]'
                                        else:
                                            expanded_value = ep.expand_path_alias(element, path_alias_sheet)
                                        
                                        # replace the element in the split path with the expanded value    
                                        split_path[index] = expanded_value
                                        
                                # re-join the split path
                                path = '.'.join(split_path)
                                
                                # add the group alias paths to the path list
                                for alias_path in group_alias_paths:
                                    with_group_alias = f'{path.split("$$")[0]}{alias_path[0]}'
                                    path_list.append([with_group_alias.split('.'), alias_path[1]])
                                                
                            # handle continuous data loop
                            elif '$loop' in path:
                                
                                # set in loop to true
                                in_loop = True
                                
                                # get csv name from path and download from GCP storage
                                csv_name = path.replace('$loop', '').strip().lower()
                                try:
                                    blob = self.bucket.blob(f'{self.continuous_data_dir}/{csv_name}')
                                except:
                                    raise self.MissingContinuousDataError(csv_name, self.project_name, personae_name, sheet)

                                csv_data_str = blob.download_as_string()
                                csv_data = csv_data_str.decode('utf-8').split('\n')
                                
                                loop_start = row[0].row
                                loop_lines = ep.get_loop_lines(scenario_sheet, loop_start, path_location[1], example_data_location[1])
                                
                                # expand loop lines
                                expanded_loop_lines = expand_loop_lines(loop_lines, csv_data, path_alias_sheet, group_alias_sheet)
                                
                                # append to path list
                                [path_list.append([arr[0].split('.'), arr[1]]) for arr in expanded_loop_lines]
                                
                            # handle path alias
                            elif '$' in path:
                                split_path = path.split('.')
                                
                                # loop over the split path if there is a $ find the expanded value and replace
                                for index, element in enumerate(split_path):
                                    if element.startswith('$'):
                                        if '[' in element:
                                            element_no_index = element.split('[')[0]
                                            element_index = element.split('[')[1].replace(']', '')
                                            expanded_value = ep.expand_path_alias(element_no_index, path_alias_sheet)
                                            expanded_value = f'{expanded_value}[{element_index}]'
                                        else:
                                            expanded_value = ep.expand_path_alias(element, path_alias_sheet)
                                            
                                        # replace the element in the split path with the expanded value
                                        split_path[index] = expanded_value
                                        
                                path_list.append([split_path, value])
                                
                            else:
                                path_list.append([path.split('.'), value])
                                
                # find event in timeline and add path data to it
                for event in storylog['timeline']:
                    if event['event'] == sheet:
                        # get pathlist in correct format
                        new_path_list = []
                        for path in path_list:
                            new_path_list.append({
                                'dataPath': path[0],
                                'exampleData': path[1]
                            })
                        
                        event['linked_data'] = new_path_list
                        
                # validate paths and make storylog
                
                for path in path_list:
                    if path[0] not in standard_data:
                        false_paths[sheet].append(path[0])
                    
        
        return false_paths, storylog
    
    def render_story(self, personae_name, personae_story_data: dict) -> None:
        '''
        A function to render a personae story
        
        :param personae_story_data: The name of the personae to render the story for
        :return: The rendered story
        '''
        # get personae story template
        template = self.template_env.get_template('story.html')
        
        # render story
        output = template.render(personae_story_data)
        
        # save to website directory on storage bucket
        blob = self.bucket.blob(f'{self.project_name}/website/Stories/{personae_name}_story.html')
        blob.upload_from_string(output)
        
        print(f"{personae_name} story rendered")
        
        return
    
    def render_log(self, personae_name, personae_timeline_data: dict) -> None:
        '''
        A function to render a personae timeline
        
        :param personae_name: The name of the personae to render the timeline for
        :param personae_timeline_data: The data of the personae to render the timeline for
        :return: None
        '''
        # get personae timeline template
        template = self.template_env.get_template('timeline.html')
        
        # render timeline
        output = template.render(personae_timeline_data)
        
        # save to website directory on storage bucket
        blob = self.bucket.blob(f'{self.project_name}/website/Timelines/{personae_name}_log.html')
        blob.upload_from_string(output)
        
        print(f"{personae_name} timeline rendered")
        
        return
        
    
    def render_data(self, personae_name, personae_data: dict, scenario_event_name) -> None:
        '''
        A function to render a personae data in the different types of style
        
        :param personae_name: The name of the personae to render the data for
        :param personae_data: The data of the personae to render the data for
        :param scenario_event_name: The name of the scenario event to render the data for
        :return: None
        '''
        # get tree view template and render
        template = self.template_env.get_template('tree_view_template.jinja')
        output = template.render(personae_data)
        blob = self.bucket.blob(f'{self.project_name}/website/data/{personae_name}/tree/{scenario_event_name}.html')
        blob.upload_from_string(output)
        
        # get rendered view template and render
        template = self.template_env.get_template('json_render.jinja')
        output = template.render(personae_data)
        blob = self.bucket.blob(f'{self.project_name}/website/data/{personae_name}/render/{scenario_event_name}.html')
        blob.upload_from_string(output)
        
        # just upload JSON data to JSON folder
        json_data = json.dumps(personae_data)
        blob = self.bucket.blob(f'{self.project_name}/website/data/{personae_name}/json/{scenario_event_name}.json')
        blob.upload_from_string(json_data)
        
        return
        

    def compile_website(self) -> None:
        '''
        A function to compile the website
        Will re-render all templates and save them to the website directory
        
        :return: None
        '''
        # get the list of personae
        personae_list = [blob.name for blob in self.bucket.list_blobs(prefix=f'{self.project_name}/personae/') if blob.name.endswith('.storylog.json')]
        
        # clear the websirte directory and re-create from website contents folder
        blobs = self.bucket.list_blobs(prefix=f'{self.project_name}/website/')
        for blob in blobs:
            blob.delete()
            
        # get the list of files in the website contents folder
        blobs = self.bucket.list_blobs(prefix=f'{self.project_name}/website_contents/')
        for blob in blobs:
            # Construct the destination object name
            destination_blob_name = blob.name.replace('website_contents', 'website', 1)

            # Copy the blob to the new location
            self.bucket.copy_blob(blob, self.bucket, destination_blob_name)
            
        # loop over personae
        for personae in personae_list:
            
            # get personae data
            blob = self.bucket.blob(personae)
            personae_data_str = blob.download_as_string()
            personae_data = json.loads(personae_data_str)
            
            # get personae name
            personae_name = personae.split('/')[-1].replace('.json', '')
            
            # get personae story
            personae_story = personae_data['story']
            
            # get personae timeline
            storylog_timeline = personae_data['timeline']
            timeline = []
            for event in storylog_timeline:
                timeline.append({
                    'time': event['time'],
                    'event': event['event'],
                    'sheet': event['sheet']
                })
            
            # render story and timeline
            self.render_story(personae_name, personae_story)
            self.render_log(personae_name, timeline) # need to get timeline in correct format or changet the template
            
            # loop over timeline
            for event in storylog_timeline:
                if event['linked_data'] != []:

                    # get the event name
                    event_name = event['event']
                    
                    # get the event data
                    event_data = event['linked_data']
                    
                    # get data in correct format
                    paths = []
                    for item in event_data:
                        paths.append(item['dataPath'].split('.'), item['exampleData'])
                        
                    # get the paths into an object
                    personae_data_object = ep.create_object(paths)
                    
                    # render data
                    self.render_data(personae_name, personae_data_object, event_name)
                
            
        return

def expand_loop_lines(loop_lines: list,
                      csv_data: list,
                      path_alias_sheet: openpyxl.worksheet.worksheet.Worksheet,
                      group_alias_sheet: openpyxl.worksheet.worksheet.Worksheet):
    '''
    A function to expand a list of loop lines
    
    :param loop_lines: The lines to expand
    :param csv_data: The csv data to loop over
    :param path_alias_sheet: The path alias sheet
    :param group_alias_sheet: The group alias sheet
    :return: The expanded loop lines
    '''
    path_list = []
    
    for index, continuous_data_arr in enumerate(csv_data):
        # loop over loop_lines
        for loop_lines_arr in loop_lines:
            path = loop_lines_arr[0]
            value = loop_lines_arr[1]
            
            # apply index to path
            path = path.replace('%', str(index))
            
            # check value isnt a reference
            if str(value).startswith('#'):
                
                # get index of value from continuous data 
                value_index = int(value.replace('#',''))
                
                # sort the csv file into array
                value = str(continuous_data_arr[value_index])
                
                # if digit then convert to int
                if value.isdigit():
                    value = int(value)
                else:
                    value = value.lower().strip()
            
            # check for group alias
            if path.split('.')[-1].startswith('$$'):
                group_alias_name = path.split('.')[-1].replace('$$', '').strip().lower()
                group_alias_paths = ep.expand_group_aliases(group_alias_sheet, group_alias_name)

                for alias_path in group_alias_paths:
                    if len(path.split('.')[:-1]) != 0:
                        path_array = '.'.join([path.split('.')[:-1], alias_path[0]]).split('.')
                    else:
                        path_array = alias_path[0].split('.')

                    path_value = alias_path[1]

                    # check for path aliases
                    for index, element in enumerate(path_array):
                        if element.startswith('$'):

                            # expand path alias
                            expanded = ep.expand_path_alias(element, path_alias_sheet)

                            # replace the element with the expanded value
                            path_array[index] = expanded

                    path_list.append(['.'.join(path_array), path_value])

            # check for path alias
            if '$' in path:
                
                # get split path
                split_path = path.split('.')
                
                # loop over the split path if there is a $ find the expanded value and replace
                for index, element in enumerate(split_path):
                    if element.startswith('$'):
                        # get the expanded value
                        expanded_value = ep.expand_path_alias(element, path_alias_sheet)
                        
                        # replace the element with the expanded value
                        split_path[index] = expanded_value

                # add to path list
                path_list.append(['.'.join(split_path), value])
            
            else:
                # add to path list
                path_list.append([path, value])
                
    return path_list