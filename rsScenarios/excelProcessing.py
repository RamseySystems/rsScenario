import os
import re
from flask import flash
import openpyxl
from rsScenarios import usefulFunctions as uf
import csv
import json
import shutil
from jinja2 import Environment, FileSystemLoader
from flask import render_template
import copy
from datetime import datetime, time, date

# custom error handler
class ExcelProcessingError(Exception):
    def __init__(self, message="Error occoured while processing excel file"):
        self.message = message
        super().__init__(self.message)
        
class MissingHeadingsError(ExcelProcessingError):
    def __init__(self, message):
        self.message = message
        super().__init__(self.message)
        
class MissingPathAliasError(ExcelProcessingError):
    def __init__(self, message="Missing path alias in excel file"):
        self.message = message
        super().__init__(self.message)

class MissingGroupAliasError(ExcelProcessingError):
    def __init__(self, message="Missing group alias in excel file"):
        self.message = message
        super().__init__(self.message)
        
class MissingSheetError(ExcelProcessingError):
    def __init__(self, message="Missing sheet in excel file"):
        super().__init__(self.message)
        
class StoryLogError(Exception):
    def __init__(self, message="Error occoured whilst processing storylog file"):
        self.message = message
        super().__init__(self.message)
        
class MissingKeysError(StoryLogError):
    def __init__(self, message="Missing keys in story log"):
        self.message = message
        super().__init__(self.message)
        
class MissingDataError(StoryLogError):
    def __init__(self, message="Empty linked data in story log"):
        self.message = message
        super().__init__(self.message)
    

class CustomEncoder(json.JSONEncoder):
    def encode_dict(self, d):
        for key, value in d.items():
            try:
                if isinstance(value, (datetime, time)):
                    d[key] = value.isoformat()
                elif isinstance(value, dict):
                    d[key] = self.encode_dict(value)
            except TypeError as e:
                print(f"Error for key '{key}': {e}")
        return d

    def encode(self, obj):
        if isinstance(obj, dict):
            obj = self.encode_dict(obj)
        return super().encode(obj)
    
def find_cell_location(sheet, cell_value: str) -> tuple:
    '''
    Find the coordinates of a cell in a worksheet and return as a tuple.

    :param sheet: The sheet to search.
    :param cell_value: The text to search for in the sheet.
    :return: A tuple containing the row and column coordinates.
    '''
    row_index = 1
    for row in sheet.iter_rows():
        cell_index = 1
        for cell in row:
            value = str(cell.value).lower().strip()
            if value == cell_value.lower():
                return row_index, cell_index
            cell_index += 1
        row_index += 1
    raise MissingHeadingsError(f"Missing headings in excel file {sheet.sheet_name}")

# not sure how to test this
def open_workbook_sheets(path: str):
    '''
    A function to load specific sheets from an excel file into an array and return required sheets.

    :param path: The path to the excel document
    :return: A tuple containing an array of sheet objects, group alias sheet, and path alias sheet
    '''
    wb = openpyxl.load_workbook(path, data_only=True)
    sheet_names = wb.sheetnames

    # Filter out unwanted sheets
    desired_sheets = [ws for ws in sheet_names if ws.lower() not in ['group aliases', 'path aliases']]
    wb_array = [wb[ws] for ws in desired_sheets]

    # Check for required sheets
    group_alias_sheet = wb['Group Aliases']
    path_alias_sheet = wb['Path Aliases']

    return wb_array, group_alias_sheet, path_alias_sheet



# a function to get the group aliases from a sheet
def expand_group_aliases(group_alias_sheet: dict, group_alias_name: str) -> list:
    '''
    A function to expand the group aliases in a sheet

    :param group_alias_sheet: The group alias sheet to process
    :param group_alias_name: The name of the group alias to expand
    :return: A list of expanded group aliases
    '''

    # get the cell locations for the group alias name and path
    alias_name_cell = find_cell_location(group_alias_sheet, 'Group Alias Name')
    alias_path_cell = find_cell_location(group_alias_sheet, 'Path')

    # get the group alias paths
    group_alias_paths = []

    for row in group_alias_sheet.iter_rows(min_row=alias_name_cell[0] + 1,
                                           max_row=group_alias_sheet.max_row,
                                           min_col=alias_name_cell[1],
                                           max_col=alias_path_cell[1]+1):
        if row[0].value:
            alias_name = row[0].value.lower().strip()
            alias_path = row[-2].value.lower().strip()
            alias_value = row[-1].value.lower().strip()

            if alias_name == group_alias_name:
                group_alias_paths.append([alias_path, alias_value])

    if group_alias_paths:
        return group_alias_paths
    return MissingGroupAliasError(f'Missing group alias {group_alias_name} in excel file')

def trim_first_elements(path_list: list):
    '''
    A function that removes the first element of each path arr in a path list

    :param path_list: the path list to process
    :return: :list:
    '''
    new_path_list = []
    for path_item in path_list:
        new_path = path_item[0][1:]
        new_path_list.append([new_path, path_item[1]])

    return new_path_list


def sort_paths(start_paths: list):
    '''
    A function that sorts the paths in a path list by their first index

    start_paths is an array of path, value
    path is an arry of path elements

    This function returns an array of arrays of path, value


    :param start_paths: the unformatted paths
    :return: :list:
    '''
    sorted_paths = []
    for path_item in start_paths:
        path = path_item[0]
        if not sorted_paths:
            sorted_paths.append([path_item])
        else:
            found = False
            for list in sorted_paths:
                key = list[0][0][0]
                if path[0] == key:
                    list.append(path_item)
                    found = True
                    break
            if not found:
                sorted_paths.append([path_item])
    return sorted_paths

# a function to turn scenario paths into a JSON object
def create_object(start_paths: list):
    '''
    A recursive function that creates a JSON object from a path list

    :param start_paths: the path list (as a list of tuples (path, value))
    :return: :dict:
    '''
    # Initialise result
    result = {}
    
    # Sort paths
    # This is an array of arrays of path, value
    # It sorts the paths by their first index
    my_paths = sort_paths(start_paths)
    
    # Loop over paths with the same first index
    for path_lists in my_paths:
        
        # Get the first element of the first path
        path_first_element = path_lists[0][0][0]
        
        # This checks for an index
        # If there is an index then it is split by that index and will therefore be an array greater than 1
        matches = re.split(r".*(\[[0-9*]*\]).*", path_first_element)
        
        # This removes the index from the path as it's only needed for sorting but not displayed
        path_first_element_without_index = path_first_element.split("[")[0]
        if len(path_lists) == 1 and len(path_lists[0][0]) == 1: # if we have a single path and value (i.e., a leaf node)    
            item_value = str(path_lists[0][1])
            
            # Check if the path has an index
            if len(matches) != 1:
                
                # Try to append to the result. If you can't, then create a new array as we have not encountered this path before
                try:
                    result[path_first_element_without_index].append(item_value)
                except:
                    result.update(
                        {path_first_element_without_index: [item_value]})

            # If there is no index then just add the value to the result
            else:
                result.update({path_first_element: item_value})
                
        # If the path is not a leaf node then we need to recurse
        else:
            
            # Check if the path has an index
            if len(matches) != 1:
                
                # Try to append to the result. If you can't, then create a new array as we have not encountered this path before
                try:
                    result[path_first_element_without_index].append(
                        create_object(trim_first_elements(path_lists)))
                except:
                    result.update({path_first_element_without_index: [create_object(trim_first_elements(path_lists))]})

            # If there is no index then just add the value to the result
            else:
                
                # If you can't trim the first element then you have reached the end of the path so return the result
                try:
                    result.update({path_first_element: create_object(
                        trim_first_elements(path_lists))})
                except:
                    return result
    return result





# a function to process the scenario sheets in an excel file
def get_scenario_path_list(scenario_sheet: dict,
                           group_alias_sheet: dict,
                           path_alias_sheet: dict,
                           continuous_data_path: list,
                           user_session: str,
                           TMP_DIR: str) -> list:
    '''
    A function to process a scenario sheet and return the paths

    :param scenario_sheet: The scenario sheet to process
    :param group_alias_sheet: The group alias sheet
    :param path_alias_sheet: The path alias sheet
    :param continuous_data_path: The path to the continuous data file
    :param TMP_DIR: The path to the temporary directory
    '''

    # get the cell locations for the path and example data
    
    path_location = find_cell_location(scenario_sheet, 'Data Path')
    example_data_location = find_cell_location(scenario_sheet, 'Example Data')


    # get the paths
    path_list = []
    duplicates = []
    in_loop = False

    for row in scenario_sheet.iter_rows(min_row=path_location[0] + 1,
                                        max_row=scenario_sheet.max_row,
                                        min_col=path_location[1],
                                        max_col=example_data_location[1]):
        
        # loop over the rows and expand the path if needed
        if row[0].value:
            path = row[0].value.lower().strip()
            value = row[1].value

            # check if in loop
            if in_loop:
                if path != '$loopend':
                    continue
                elif path == '$loopend':
                    in_loop = False
                    continue

            # process group alias
            elif '$$' in path:
                
                # get the group alias name
                group_alias_name = path.split('.')[-1].replace('$$', '').strip().lower()

                # get the rest of the path
                without_alias = '.'.join(path.split('.')[:-1])

                # get the group alias paths
                group_alias_paths = expand_group_aliases(group_alias_sheet, group_alias_name)

                split_path = path.split('.')
                for index, element in enumerate(split_path[:-1]):
                    if element.startswith('$'):
                        # if element has an index, remove it
                        if '[' in element:
                            element_no_index = element.split('[')[0]
                            element_index = element.split('[')[1].replace(']', '')
                        
                            # get the expanded value
                            expanded_value = expand_path_alias(element_no_index, path_alias_sheet)
                            expanded_value = f'{expanded_value}[{element_index}]'
                        else:
                            # get the expanded value
                            expanded_value = expand_path_alias(element, path_alias_sheet)
                        
                        # replace the element with the expanded value
                        split_path[index] = expanded_value

                path = '.'.join(split_path)

                for alias_path in group_alias_paths:
                    with_group_alias = f'{path.split("$$")[0]}{alias_path[0]}'
                    path_list.append([with_group_alias.split('.'), alias_path[1]])

            elif '$loop' in path:
                
                # set in loop to true
                in_loop = True
                
                # get csv name & path
                csv_name = path.replace('$loop', '').strip().lower()
                continuous_data_path = f'{TMP_DIR}/{user_session}/upload/{csv_name}'
                
                # get loop lines
                loop_start = row[0].row
                
                # get loop lines
                loop_lines = get_loop_lines(scenario_sheet, loop_start, path_location[1], example_data_location[1])
                
                # expand the loop lines
                expanded_loop_lines = expand_loop_lines(loop_lines, continuous_data_path, path_alias_sheet, group_alias_sheet)

                # append to path list
                [path_list.append([arr[0].split('.'), arr[1]]) for arr in expanded_loop_lines]
                
            elif '$' in path:
                
                # get split path
                split_path = path.split('.')
                
                # loop over the split path if there is a $ find the expanded value and replace
                for index, element in enumerate(split_path):
                    if element.startswith('$'):
                        # if element has an index, remove it
                        if '[' in element:
                            element_no_index = element.split('[')[0]
                            element_index = element.split('[')[1].replace(']', '')
                        
                            # get the expanded value
                            expanded_value = expand_path_alias(element_no_index, path_alias_sheet)
                            expanded_value = f'{expanded_value}[{element_index}]'
                        else:
                            # get the expanded value
                            expanded_value = expand_path_alias(element, path_alias_sheet)
                        
                        # replace the element with the expanded value
                        split_path[index] = expanded_value

                # add to path list
                path_list.append([split_path, value])
            
            else:
                # add to path list
                path_list.append([path.split('.'), value])
                
            if [path.split('.'), value] in path_list[:-1]:
                duplicates.append(path_list[-1])

    return path_list, duplicates
                

            


# a function to expand path aliases
def expand_path_alias(path_alias_name, path_alias_sheet):
    '''
    A function to expand path aliases
    :param path_alias_name: The name of the path alias to expand
    :param path_alias_sheet: The path alias sheet
    
    :return: A list of expanded path aliases
    '''
    
    # get the cell locations for the path alias name and path
    alias_name_cell = find_cell_location(path_alias_sheet, 'Path Alias Name')
    alias_path_cell = find_cell_location(path_alias_sheet, 'Value')
    
    origional_name = path_alias_name
    
    if '[' in path_alias_name:
        path_alias_name = origional_name.split('[')[0]
        alias_index = origional_name.split('[')[-1][:-1]
    else:
        alias_index = None

    # get the path alias paths
    for row in path_alias_sheet.iter_rows(min_row=alias_name_cell[0] + 1,
                                          max_row=path_alias_sheet.max_row,
                                          min_col=alias_name_cell[1],
                                          max_col=alias_path_cell[1]+1):
        alias_name = row[0].value.lower().strip()
        alias_value = row[1].value.lower().strip()
        
        if alias_value and alias_name:
            if alias_name == path_alias_name:
                if '$' in alias_value:
                    # split the path alias value
                    split_value = alias_value.split('.')
                    
                    # loop over the split value if there is a $ find the expanded value and replace
                    for index, element in enumerate(split_value):
                        if element.startswith('$'):
                            # get the expanded value
                            expanded_value = expand_path_alias(element, path_alias_sheet)
                            
                            # replace the element with the expanded value
                            split_value[index] = expanded_value
                            
                            if alias_index != None:
                                return f'{".".join(split_value)}[{alias_index}]'
                            else: 
                                return '.'.join(split_value)
                else:
                    if alias_index != None:
                        return f'{alias_value}[{alias_index}]'
                    else:
                        return alias_value
    raise MissingPathAliasError(f'Missing path alias {path_alias_name} in excel file')
    

def get_loop_lines(scenario_sheet: dict, loop_start: int, path_cell: int, value_cell: int) -> list:
    '''
    A function to get the loop lines from a scenario sheet
    
    :param scenario_sheet: The scenario sheet to process
    :param loop_start: The row number of the loop start
    :param path_cell: The column number of the path cell
    :param value_cell: The column number of the value cell
    :return: :list:
    '''
    
    loop_lines = []
    for row in scenario_sheet.iter_rows(min_row=loop_start+1,
                                        max_row=scenario_sheet.max_row,
                                        min_col=path_cell,
                                        max_col=value_cell):
        if row[0].value:
            path = row[0].value.lower().strip()
            value = row[1].value
            
            if path == '$loopend':
                return loop_lines
            
            loop_lines.append([path, value])

def expand_loop_lines(loop_lines: list, 
                      continuous_data_path: str,
                      path_alias_sheet: dict,
                      group_alias_sheet: dict):
    '''
    A function to expand the loop lines
    
    :param loop_lines: The loop lines to expand
    :param continuous_data_path: The path to the continuous data file
    :param path_alias_sheet: The path alias sheet
    :param group_alias_sheet: The group alias sheet
    :return: :list:
    '''
    
    path_list = []
    
    # open continuous data file
    with open(continuous_data_path) as csv_file:
        continuous_data = list(csv.reader(csv_file, delimiter=','))
        
    for index, continuous_data_arr in enumerate(continuous_data):


        # loop over loop_lines
        for loop_lines_arr in loop_lines:
            path = loop_lines_arr[0]
            value = loop_lines_arr[1]
            
            # apply index to path
            path = path.replace('%', str(index))
            
            # check value isnt a reference
            if str(value).startswith('#'):
                
                # get index of value from continuous data 
                value_index = int(value.replace('#',''))
                
                # sort the csv file into array
                # test
                
                value = str(continuous_data_arr[value_index])
                
                
                # if digit then convert to int
                if value.isdigit():
                    value = int(value)
                else:
                    value = value.lower().strip()
            
            # check for group alias
            if path.split('.')[-1].startswith('$$'):
                group_alias_name = path.split('.')[-1].replace('$$', '').strip().lower()
                group_alias_paths = expand_group_aliases(group_alias_sheet, group_alias_name)

                for alias_path in group_alias_paths:
                    if len(path.split('.')[:-1]) != 0:
                        path_array = '.'.join([path.split('.')[:-1], alias_path[0]]).split('.')
                    else:
                        path_array = alias_path[0].split('.')

                    path_value = alias_path[1]

                    # check for path aliases
                    for index, element in enumerate(path_array):
                        if element.startswith('$'):

                            # expand path alias
                            expanded = expand_path_alias(element, path_alias_sheet)

                            # replace the element with the expanded value
                            path_array[index] = expanded

                    path_list.append(['.'.join(path_array), path_value])

            # check for path alias
            if '$' in path:
                
                # get split path
                split_path = path.split('.')
                
                # loop over the split path if there is a $ find the expanded value and replace
                for index, element in enumerate(split_path):
                    if element.startswith('$'):
                        # get the expanded value
                        expanded_value = expand_path_alias(element, path_alias_sheet)
                        
                        # replace the element with the expanded value
                        split_path[index] = expanded_value

                # add to path list
                path_list.append(['.'.join(split_path), value])
            
            else:
                # add to path list
                path_list.append([path, value])
                
    return path_list


def serialize_datetime(obj): 
    if isinstance(obj, datetime) or isinstance(obj, time) or isinstance(obj, date): 
        return obj.isoformat() 
    raise TypeError("Type not serializable")


# a main wrapper function for the scenario tool
def scenario_tool(upload_dir: str, 
                  user_session: str, 
                  TMP_DIR: str, 
                  CURRENT_DIR: str,
                  scenario_files: list,
                  standard_paths: list,
                  continuous_data_files: list,
                  personae: str) -> dict:
    '''
    A function that processes the scenario files and returns a website folder displaying the rendered output
    
    :param upload_dir: The upload directory
    :param user_session: The user session
    :param TMP_DIR: The temporary directory
    :param CURRENT_DIR: The current directory
    :param scenario_files: The scenario files
    :param standard_paths: The standard paths
    :param continuous_data_files: The continuous data files
    :param personae: The personae
    :return: dict : The summary object containing false paths
    '''
    try:
        os.mkdir(f'{TMP_DIR}/{user_session}/website')
    except FileExistsError:
        shutil.rmtree(f'{TMP_DIR}/{user_session}/website')
        os.mkdir(f'{TMP_DIR}/{user_session}/website')

    website_contents_path = f'{CURRENT_DIR}/renders_website_content'
    uf.clear_and_replace(website_contents_path, f'{TMP_DIR}/{user_session}/website')
        
    # builds personae list
    personae = []
    for file in os.listdir(upload_dir):
        if file.split('.')[0].endswith('storylog'):
            personae.append(file.replace('-storylog', ''))
        elif file.split('.')[-1] != 'json' and file[0] != '_' and file != "Provenance.xlsx":
            personae.append(file)
        
     # clear story & log directorys and remake
    try:     
        shutil.rmtree(f'{TMP_DIR}/{user_session}/website/Stories')
    except:
        pass
    os.mkdir(f'{TMP_DIR}/{user_session}/website/Stories')
    
    try:
        shutil.rmtree(f'{TMP_DIR}/{user_session}/website/logs')
    except:
        pass
    os.mkdir(f'{TMP_DIR}/{user_session}/website/logs')
    
    # render the index tempalte
    env = Environment(loader=FileSystemLoader('templates'))
    template = env.get_template('index_template.jinja')
    output = template.render(personae=personae)

    with open(f'{TMP_DIR}/{user_session}/website/index.html', 'w') as f:
        f.write(output)
    
    # get the scenario paths
    scenario_false_paths = {}
    scenario_duplicates = {}
    exported_model = {}
    for scenario in scenario_files:
        file_name = scenario.split('.')[0]
        
        # create storage locations
        os.mkdir(f'{TMP_DIR}/{user_session}/website/data/{file_name}')
        os.mkdir(f'{TMP_DIR}/{user_session}/website/data/{file_name}/json/')
        os.mkdir(f'{TMP_DIR}/{user_session}/website/data/{file_name}/rendered/')
        os.mkdir(f'{TMP_DIR}/{user_session}/website/data/{file_name}/tree/')
        
        
        # Get sheets from scenario file
        scenario_path = f'{upload_dir}/{scenario}'
        wb = openpyxl.load_workbook(scenario_path, data_only=True)
        sheets = wb.sheetnames
        
        scenario_sheets = []
        
        # loop over the sheet names
        for sheet in sheets:
            if sheet.lower() == 'group aliases':
                group_alias_sheet = wb[sheet]
            elif sheet.lower() == 'path aliases':
                path_alias_sheet = wb[sheet]
            elif sheet.lower() == 'story':
                story_sheet = wb[sheet]
            elif sheet.lower() == 'time line':
                time_line_sheet = wb[sheet]
            else:
                scenario_sheets.append(wb[sheet])
       
        # story
        story_object = render_story_object(story_sheet)
        uf.render_template('story.html', story_object, f'{TMP_DIR}/{user_session}/website/Stories/{file_name}_story.html')
        
        # timeline
        time_line_object = render_timeline_object(time_line_sheet, sheets, file_name)
        uf.render_template('log.html', time_line_object, f'{TMP_DIR}/{user_session}/website/logs/{file_name}_log.html', name=file_name)
        
        combined_timeline_story = {
            'summary': story_object['summary'],
            'rationale': story_object['rationale'],
            'story': story_object['story'],
            'standard_url': story_object['standard_url'],
            'standard_name': story_object['standard_name'],
            'timeline': time_line_object
        }
        
        scenario_false_paths[file_name] = {}
        scenario_duplicates[file_name] = {}
        
        for scenario in scenario_sheets:
            scenario_paths, duplicates = get_scenario_path_list(scenario, group_alias_sheet, path_alias_sheet, continuous_data_files, user_session,TMP_DIR)
            
            if file_name != 'Provenance.xlsx':
                # save scenario paths
                with open(f'{TMP_DIR}/{user_session}/output/logs/{file_name}-{scenario.title}.paths.json', 'w') as f:
                    json.dump(scenario_paths, f, indent=4, cls=CustomEncoder, default=serialize_datetime)

            # check for invalid paths
            false_paths = validate_scenario_paths(scenario_paths, standard_paths)
            scenario_name = scenario.title
            scenario_false_paths[file_name][scenario_name] = false_paths
            scenario_duplicates[file_name][scenario_name] = duplicates
            
            # make the scenario paths into a JSON object
            scenario_object = create_object(scenario_paths)
            
            # render the scenario object
            uf.render_template('json_render.html', scenario_object, f'{TMP_DIR}/{user_session}/website/data/{file_name}/rendered/{scenario_name}.html')
            uf.render_template('tree_view_template.jinja', scenario_object, f'{TMP_DIR}/{user_session}/website/data/{file_name}/tree/{scenario_name}.html')
            with open(f'{TMP_DIR}/{user_session}/website/data/{file_name}/json/{scenario_name}.json', 'w') as f:
                json.dump(scenario_object, f, indent=4)
                
            for index, event in enumerate(combined_timeline_story['timeline']):
                sheet = event['sheet']
                event['standard'] = story_object['standard_name']
                if sheet == scenario_name:
                    combined_timeline_story['timeline'][index]['linked_data'] = convert_paths_to_linked_data(scenario_paths)
                    break
        try:
            with open(f'{TMP_DIR}/{user_session}/output/logs/{file_name}-storylog.json', 'w') as f:
                json.dump(combined_timeline_story, f, indent=4, cls=CustomEncoder, default=serialize_datetime)
        except TypeError as e:
            print(e)
            raise TypeError(e)
    
    return scenario_duplicates, scenario_false_paths
    
    
def convert_paths_to_linked_data(scenario_paths: list):
    converted = []
    for arr in scenario_paths:
        converted.append({
            'dataPath': '.'.join(arr[0]),
            'exampleData': arr[1]
        })
        
    return converted


def render_story_object(story_sheet: dict) -> dict:
    '''
    A function to render the story sheet
    
    :param story_sheet: The story sheet to process
    :return: :dict:
    '''
    summary_cell = find_cell_location(story_sheet, 'Summary')
    rationale_cell = find_cell_location(story_sheet, 'Rationale')
    story_cell = find_cell_location(story_sheet, 'Story')
    standard_url_cell = find_cell_location(story_sheet, 'Standard URL')
    standard_name_cell = find_cell_location(story_sheet, 'Standard Name')

    story = {}
    story['summary'] = story_sheet.cell(
        row=summary_cell[0]+1, column=summary_cell[1]).value
    story['rationale'] = story_sheet.cell(
        row=rationale_cell[0]+1, column=rationale_cell[1]).value
    story['story'] = story_sheet.cell(
        row=story_cell[0]+1, column=story_cell[1]).value
    story['standard_url'] = story_sheet.cell(
        row=standard_url_cell[0]+1, column=standard_url_cell[1]).value
    story['standard_name'] = story_sheet.cell(
        row=standard_name_cell[0]+1, column=standard_name_cell[1]).value
    
    return story

def render_timeline_object(time_line_sheet: dict, sheet_names: list, file_name: str) -> dict:
    '''
    A function to render the timeline sheet
    
    :param time_line_sheet: The timeline sheet to process
    :param sheet_names: The names of the sheets
    :return: :dict:
    '''
    time_cell = find_cell_location(time_line_sheet, 'Date/Time')
    event_cell = find_cell_location(time_line_sheet, 'Event')
    sheet_cell = find_cell_location(time_line_sheet, 'Sheet')

    time_line = []
    false_sheets = []

    for row in time_line_sheet.iter_rows(time_cell[0]+1, time_line_sheet.max_row, time_cell[1], sheet_cell[1]):
        time = str(row[0].value)
        event = str(row[1].value)
        work_sheet = str(row[2].value)

        if str(work_sheet) == 'None':
            work_sheet = ''

        if not work_sheet.strip() == '':
            if work_sheet not in sheet_names:
                print(f'{work_sheet} does not exist!')
                false_sheets.append(work_sheet)

        if str(time) != 'None':

            time_line.append({
                'time': time,
                'event': event,
                'sheet': work_sheet
            })
            
    if false_sheets:
        raise ValueError(f'Sheets do not exist in {file_name}: {false_sheets}')
    
    return time_line



            
def get_standard_paths(standard_path: str, provenance_paths: list, for_export: bool = False):
    '''
    This is a function that extracts the paths from a standard in FHIR shorthand format

    :param standard_path: the path to the standard
    :return: :list:
    '''
    # load standard
    wb = openpyxl.load_workbook(standard_path, data_only=True)
    sheets = wb.sheetnames
    ws = wb[sheets[0]]

    # get row stard and path column
    name_title_cell = find_cell_location(ws, 'Name')
    info_type_cell = find_cell_location(ws, 'Information Type')

    # get single_indent
    single_indent = get_whitespace(
        str(ws.cell(row=name_title_cell[0]+2, column=name_title_cell[1]).value))

    # make path list
    path_list = []
    path = []
    previous_indent = 0
    for row in ws.iter_rows(name_title_cell[0]+1, ws.max_row, name_title_cell[1], info_type_cell[1]):
        if row:
            line = str(row[0].value)
            if line != 'None':
                info_type = str(row[-1].value)
                cardinality = str(row[2].value)
                line_indent = int(get_whitespace(line) / single_indent)
                stripped = re.sub(
                    re.escape('\\xa0\\xa0\\xa0\\xa0'), r'', line).strip().lower()
                path = get_path(previous_indent, line_indent, path, stripped)
                str_path = '.'.join(path.copy()).lower().replace('\'', '').lower()

                if provenance_paths:
                    with_provenance = []
                    if info_type == 'Event.Record':
                        for line in provenance_paths:
                            if len(line) > 1:
                                if line[0] == 'event record':
                                    provenance_line = '.'.join(line[1:])
                                    extended_path = f'{str_path}.{provenance_line}'.lower()
                                    with_provenance.append(extended_path)
                    elif info_type == 'Record':
                        for line in provenance_paths:
                            if len(line) > 1:
                                if line[0] == 'record':
                                    provenance_line = '.'.join(line[1:])
                                    extended_path = f'{str_path}.{provenance_line}'.lower()
                                    with_provenance.append(extended_path)

                    [path_list.append(line) for line in with_provenance]
                else:
                    path_list.append(path.copy())
                    
                previous_indent = line_indent
                
                
                
                if for_export:
                    if cardinality == '0...*' or cardinality == '1...*':
                        path_list.append(f'{str_path}[]')
                    else:    
                        path_list.append(str_path)

    return path_list


def export_standard(standard_path: str, 
                    provenance_path: str):
    
    # go over the provenance files
    wb = openpyxl.load_workbook(provenance_path, read_only=True)
    sheets = wb.sheetnames
    ws = wb[sheets[0]]
    
    # get row stard and path column
    name_title_cell = find_cell_location(ws, 'Name')
    info_type_cell = find_cell_location(ws, 'Information Type')

    # get single_indent
    single_indent = get_whitespace(
        str(ws.cell(row=name_title_cell[0]+2, column=name_title_cell[1]).value))
    
    provenance_paths = []
    repeating_paths = []
    path = []
    previous_indent = 0
    
    for row in ws.iter_rows(name_title_cell[0]+1, 
                            ws.max_row,
                            name_title_cell[1], 
                            info_type_cell[1]):
        if row:
            line = str(row[0].value)
            if line != 'None':
                cardinality = str(row[2].value)
                line_indent = int(get_whitespace(line) / single_indent)
                stripped = re.sub(
                    re.escape('\\xa0\\xa0\\xa0\\xa0'), r'', line).strip()
                path = get_path(previous_indent, line_indent, path, stripped)
                dot_path = '.'.join(path.copy()).lower().replace('\'', '')
                
                if cardinality == '0...*' or cardinality == '1...*':
                    repeating_paths.append(dot_path)
                    
                    real_path = find_lists(dot_path, repeating_paths)
                    provenance_paths.append(real_path)
                else:
                    real_path = find_lists(dot_path, repeating_paths)
                    provenance_paths.append(real_path)
                    
                previous_indent = line_indent
                
                
                
    # get standard paths
    wb = openpyxl.load_workbook(standard_path, data_only=True)
    sheets = wb.sheetnames
    ws = wb[sheets[0]]
    
    # get row stard and path column
    name_title_cell = find_cell_location(ws, 'Name')
    info_type_cell = find_cell_location(ws, 'Information Type')

    # get single_indent
    single_indent = get_whitespace(
        str(ws.cell(row=name_title_cell[0]+2, column=name_title_cell[1]).value))
    
    standard_paths =[]
    repeating_paths = []
    path = []
    previous_indent = 0
    for row in ws.iter_rows(name_title_cell[0]+1,
                            ws.max_row,
                            name_title_cell[1],
                            info_type_cell[1]):
        if row:
            line = str(row[0].value)
            if line != 'None':
                info_type = str(row[-1].value)
                cardinality = str(row[2].value)
                line_indent = int(get_whitespace(line) / single_indent)
                stripped = re.sub(
                    re.escape('\\xa0\\xa0\\xa0\\xa0'), r'', line).strip()
                path = get_path(previous_indent, line_indent, path, stripped)
                dot_path = '.'.join(path.copy()).lower().replace('\'', '').lower()
                
                if cardinality == '0...*' or cardinality == '1...*':
                    repeating_paths.append(dot_path)
                    real_path = find_lists(dot_path, repeating_paths)
                else:
                    real_path = find_lists(dot_path, repeating_paths)

                with_provenance = []
                if info_type == 'Event.Record':
                    for line in provenance_paths:
                        if line.split('.')[0] == 'event record':
                            provenance_line = '.'.join(line.split('.')[1:])
                            extended_path = f'{real_path}.{provenance_line}'
                            with_provenance.append(extended_path.lower())
                elif info_type == 'Record':
                    for line in provenance_paths:
                        if line.split('.')[0] == 'record':
                            provenance_line = '.'.join(line.split('.')[1:])
                            extended_path = f'{real_path}.{provenance_line}'
                            with_provenance.append(extended_path.lower())

                [standard_paths.append(line) for line in with_provenance]
                
                previous_indent = line_indent

    return standard_paths
            
    
def find_lists(original_path, list_of_paths):
    original_path_arr = original_path.split('.')
    
    for path in list_of_paths:
        for index, element in enumerate(original_path_arr):
            if element == path:
                original_path_arr[index] = f'{element}[]'

    return '.'.join(original_path_arr).lower()


def get_whitespace(line: str):
    '''
    A functio to return the amount of white space at the start of a string

    :param line: the line to process
    :return: :int:
    '''
    full_len = len(line)
    stripped = len(line.lstrip())
    whitespace = full_len - stripped

    return whitespace


def get_path(previous_indent: int, indent: int, path: list, stripped: str):
    '''
    A function to get the current data path

    :param previous_indent: the last lines indentation level
    :param indent: the current lines indentation level
    :param path; the previous path
    :param stripped: the data element string stripped of leading and trailing whitespace
    :return: :list:
    '''
    # Get path for A
    if indent == 0:  # No indent
        path = []
        path.append(stripped)
    elif previous_indent == indent:  # Same indent
        path.pop()
        path.append(stripped)
    elif previous_indent < indent:  # More indented
        path.append(stripped)
    elif previous_indent > indent:  # Less indented
        num_less = previous_indent - indent + 1
        path = path[:int(len(path)-num_less)]
        path.append(stripped)
    return path


# a function to validate the scenario paths
def validate_scenario_paths(scenario_paths: list, standard_paths: list) -> list:
    '''
    A function to validate the scenario paths
    Returns a list of false paths
    
    :param scenario_paths: The scenario paths to validate
    :param standard_paths: The standard paths to validate against
    '''
    
    # lowercase the standard paths
    for index, element in enumerate(standard_paths):
        standard_paths[index] = element.lower()
    
    
    false_paths = []
    for scenario_path in scenario_paths:
        if isinstance(scenario_path[0], str):
            joined = scenario_path[0]
        else:
            joined = '.'.join(scenario_path[0])
        no_alias = re.sub(r'\[\d+\]', '', joined)
        split_no_alias = no_alias.split('.')  
        if no_alias not in standard_paths:
            false_paths.append([split_no_alias, scenario_path[1]])
            
    return false_paths

def validate_storylog_paths(scenario_paths: list, standard_paths: list) -> list:
    '''
    A function to validate the paths of a storylog
    
    :param scenario_paths: The scenario paths to validate
    :param standard_paths: The standard paths to validate against
    '''
    # lowercase the standard paths
    for index, element in enumerate(standard_paths):
        standard_paths[index] = element.lower()
        
    false_paths = []
    for element in scenario_paths:
        path = element['dataPath']
        no_alias = re.sub(r'\[\d+\]', '', path)
        if no_alias not in standard_paths:
            false_paths.append(path)
            
    return false_paths
    

# a function to retrieve the parts from the story log in the format for the templates
def extract_data_from_storylog(story_log: dict):
    
    scenarios = {}
    timeline = []
    story = {
        "summary": story_log["summary"],
        "rationale": story_log["rationale"],
        "story": story_log["story"],
        "standard_url": story_log["standard_url"],
        "standard_name": story_log["standard_name"]
    }
    
    # get the scenarios data
    for event in story_log['timeline']:
        if 'linked_data' in event.keys():
            if len(event['linked_data']) > 0:
                raise  MissingDataError(f'Linked data is empty in {event["sheet"]}')

            scenarios[event['sheet']] = event['linked_data']
        timeline.append(
            {
                'time': event['time'],
                'event': event['event'],
                'sheet': event['sheet']
            }
        )
    
    return story, timeline, scenarios
    
def get_duplicates(scenario_paths: list) -> list:
    duplicates = []
    scenario_path_list = []
    for path in scenario_paths:
        if path in scenario_path_list:
            duplicates.append(path)
        scenario_path_list.append(path)
        
    return duplicates

def get_timeline_storylog_format(timeline_sheet, standard_name: str) -> dict:
    '''
    A function to get the timeline in the format for the storylog
    
    :param timeline_sheet: The timeline sheet to process
    :return: :dict:
    '''
    timeline = []
    
    # get the cell locations for the path and example data
    time_cell = find_cell_location(timeline_sheet, 'Date/Time')
    event_cell = find_cell_location(timeline_sheet, 'Event')
    sheet_cell = find_cell_location(timeline_sheet, 'Sheet')
    
    for row in timeline_sheet.iter_rows(time_cell[0]+1, timeline_sheet.max_row, time_cell[1], sheet_cell[1]):
        time = str(row[0].value)
        event = str(row[1].value)
        work_sheet = str(row[2].value)
        
        if str(work_sheet) == 'None':
            work_sheet = ''
        
        if str(time) != 'None':
            timeline.append({
                'time': time,
                'event': event,
                'sheet': work_sheet,
                'standard': standard_name,
                'linked_data': []
            })
            
    return timeline
            